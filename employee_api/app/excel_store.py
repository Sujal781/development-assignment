"""
app/excel_store.py

Handles all reading and writing of employee records to the Excel
file that acts as this API's database.

Single Responsibility: this class only knows how to translate
between rows in employees.xlsx and plain employee dicts. It has no
idea HTTP, FastAPI, or external APIs exist - the same separation of
concerns used by FileReader/FileWriter in the customer demo.
"""

import os
from threading import Lock
from typing import Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# Column order in the sheet. employee_id is always column A.
COLUMNS = ["employee_id", "name", "email", "department", "designation", "salary"]
EMAIL_COL = COLUMNS.index("email")


class ExcelEmployeeStore:
    """Reads and writes employee records against a single Excel file."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        # Guards against two requests writing to the file at the same
        # instant. This only protects against concurrent requests within
        # this one process - see the README for why that's a real
        # limitation of using a flat file as a "database".
        self._lock = Lock()
        self._ensure_file_exists()

    def _ensure_file_exists(self) -> None:
        """Create the Excel file with just a header row if it doesn't exist yet."""
        if os.path.exists(self.file_path):
            return
        folder = os.path.dirname(self.file_path)
        if folder:
            os.makedirs(folder, exist_ok=True)

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Employees"
        sheet.append(COLUMNS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        wb.save(self.file_path)

    def _load_sheet(self):
        wb = load_workbook(self.file_path)
        return wb, wb["Employees"]

    def get_all(self) -> list:
        """Return every employee record as a list of dicts."""
        _, sheet = self._load_sheet()
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue  # skip any blank trailing rows
            records.append(dict(zip(COLUMNS, row)))
        return records

    def get_by_id(self, employee_id: str) -> Optional[dict]:
        """Return one employee record, or None if the ID doesn't exist."""
        for record in self.get_all():
            if record["employee_id"] == employee_id:
                return record
        return None

    def _next_id(self, sheet) -> str:
        """Scan existing IDs and return the next sequential EMP#### ID."""
        highest = 1000
        for row in sheet.iter_rows(min_row=2, values_only=True):
            emp_id = row[0]
            if emp_id and str(emp_id).startswith("EMP"):
                try:
                    highest = max(highest, int(str(emp_id)[3:]))
                except ValueError:
                    continue
        return f"EMP{highest + 1}"

    def create(self, data: dict) -> dict:
        """Append a new employee record and return it (with its new ID)."""
        with self._lock:
            wb, sheet = self._load_sheet()
            new_id = self._next_id(sheet)
            record = {"employee_id": new_id, **data}
            sheet.append([record[col] for col in COLUMNS])
            wb.save(self.file_path)
            return record

    def update(self, employee_id: str, data: dict) -> Optional[dict]:
        """Update only the provided fields on an existing record.
        Returns the updated record, or None if employee_id doesn't exist."""
        with self._lock:
            wb, sheet = self._load_sheet()
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == employee_id:
                    for key, value in data.items():
                        if value is None:
                            continue
                        row[COLUMNS.index(key)].value = value
                    wb.save(self.file_path)
                    return dict(zip(COLUMNS, [c.value for c in row]))
            return None

    def delete(self, employee_id: str) -> bool:
        """Delete a record by ID. Returns True if something was deleted."""
        with self._lock:
            wb, sheet = self._load_sheet()
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == employee_id:
                    sheet.delete_rows(row[0].row, 1)
                    wb.save(self.file_path)
                    return True
            return False

    def upsert_by_email(self, data: dict) -> Tuple[dict, bool]:
        """
        Used by the external-sync flow: this mirrors the GCID demo's
        identity-resolution idea, just keyed on email instead of email-
        to-GCID. If a record with this email already exists, update it
        in place and reuse its employee_id. Otherwise, create a new
        record with a freshly generated employee_id.

        Returns (record, was_created).
        """
        with self._lock:
            wb, sheet = self._load_sheet()
            for row in sheet.iter_rows(min_row=2):
                if row[EMAIL_COL].value == data["email"]:
                    for key, value in data.items():
                        if key == "employee_id":
                            continue
                        row[COLUMNS.index(key)].value = value
                    wb.save(self.file_path)
                    return dict(zip(COLUMNS, [c.value for c in row])), False

            new_id = self._next_id(sheet)
            record = {"employee_id": new_id, **data}
            sheet.append([record[col] for col in COLUMNS])
            wb.save(self.file_path)
            return record, True
